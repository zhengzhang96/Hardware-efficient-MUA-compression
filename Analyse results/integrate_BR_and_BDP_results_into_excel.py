# -*- coding: utf-8 -*-
"""
Created on Mon Nov 22 16:24:01 2021

@author: ows18

Script for formatting the BR and BDP results as columns that can be copy-pasted
into the excel spreadhseet with the combined results.
NOTE: Excel spreadsheet musty be closed to update it.
"""
import numpy as np
import pickle
import re
from openpyxl import load_workbook

################### SELECT OPTION, PARAMETERS ############################

# Parameters        
# NOTE: bits_per_channel_for_histogram_vector HAS TO BE THE SAME AS IN THE 
# "get_BR_no_sort" AND "get_BR_with_approx_sort" SCRIPTS.
bits_per_channel_for_histogram_vector = pow(2,np.array([2,3,4,5,6,7,8,9,10]))
CV_vector = np.arange(1,7,1)
bin_vector = [1,5,10,20,50,100]
S_vector = np.arange(2,11,1)

# Specify root directory (where directories.txt file is located)
root_directory = r'D:\Dropbox (Imperial NGNI)\NGNI Share\Workspace\Oscar\Work\MUA compression\Upload code'

##########################################################################



# Read directories.txt file
with open(root_directory + '\\directories.txt') as f:
    lines = f.readlines()

# Get path to BDP results
for path in lines:
    if path.startswith('BDP_results'):
        pattern = "'(.*?)'"
        BDP_results_directory = re.search(pattern, path).group(1)
        
            
# Get excel spreadsheet directory (has the hardware processing power and 
# resources results, in this script we we will add the BR and BDP results)
for path in lines:
    if path.startswith('combined_results_excel_path'):
        pattern = "'(.*?)'"
        excel_directory = re.search(pattern, path).group(1)
        excel_path = excel_directory + '\combined_results_template.xlsx'
     
# Get SCLV directory
for path in lines:
    if path.startswith('SCLV_path'):
        pattern = "'(.*?)'"
        SCLV_directory = re.search(pattern, path).group(1)
        

# Load excel spreadsheet template, from it we get the standard format for the results
workbook = load_workbook(filename=excel_path)
sheet = workbook.active

# First and last rows in excel spreadsheet
first_row = 3
last_row = 275

excel_parameter_format = []
for value in sheet.iter_rows(min_row=first_row,max_row=last_row,min_col=1,max_col=4,values_only=True):
    excel_parameter_format.append(value)
excel_parameter_format = np.array(excel_parameter_format)


""" Load BR results """ 
# Iterate tghrough the approx. and no sort BR results
for approx_or_no_sort in ['approx-sort','no-sort']:
    
    # Get BR results directory
    if approx_or_no_sort == 'no-sort':
        for path in lines:
            if path.startswith('BR_no_sort_results'):
                pattern = "'(.*?)'"
                BR_results_directory = re.search(pattern, path).group(1)
                
    elif approx_or_no_sort == 'approx-sort': 
        for path in lines:
            if path.startswith('BR_approx_sort_results'):
                pattern = "'(.*?)'"
                BR_results_directory = re.search(pattern, path).group(1)   
                
                
    # Load results across cross-validation runs (CVs)
    cv_count = 0
    for CV in CV_vector:
        cv_count += 1
        formatted_results = []
        
        # Iterate through binning periods
        for BP in bin_vector:
        
           # Iterate through S values (MUA dynamic range)
            for S in S_vector:
                
                # BR results
                file_name = BR_results_directory + '\\BRs_S_' + \
                str(S) + '_BP_' + str(BP)+'_CV_'+str(CV)+'.pkl'
                
                # Load BR results
                with open(file_name, 'rb') as file:
                # Call load method to deserialze
                    results = pickle.load(file)
                    stored_all_var_BRs = results.get('stored_all_var_BRs') # BR results
                    encoder_red_rounds = len(stored_all_var_BRs) # total number of trianing/validation (i.e. SCLV reduction) rounds
    
                # Iterate through encoder reduction rounds
                for encoder_index, encoder_res in enumerate(stored_all_var_BRs): 
                
                    for hist_id, hist_res in enumerate(encoder_res):
    
                        # Average BR across all channels
                        mean_res = np.mean(np.array(hist_res)) 
                        worst_channel_res = np.max(np.array(hist_res)) # worst BR across all channels
                        
                        # Convert histogram size to bits/bin, gives a nice linear scale, e.g. for plots
                        histogram_log_2 = int(np.log2(bits_per_channel_for_histogram_vector[hist_id]))
                        
                        # Format results (encoder_red_rounds-encoder_index == how mnay encoders used in that round)
                        res_vec = [BP, S, histogram_log_2, encoder_red_rounds-encoder_index,\
                                   mean_res] # we store all indices and the result next to each other
                            
                        # Store all results in 2D matrix
                        formatted_results.append(res_vec)
                  
        # Store results across cross-validation runs
        if cv_count == 1:
            mean_across_CVs_formatted_results = np.array(formatted_results)
        else:                   
            mean_across_CVs_formatted_results += np.array(formatted_results)         
        
    # Given as BR, average across all CVs
    mean_across_CVs_formatted_results = mean_across_CVs_formatted_results / cv_count          
    
    
    # Iterate through excel spreadsheet, match excel rows with "formatted_results" rows.
    BR_vector = np.zeros((len(excel_parameter_format[:,0])))
    for row_ind in np.arange(len(excel_parameter_format[:,0])):
    
        a = np.where(mean_across_CVs_formatted_results[:,0] == excel_parameter_format[row_ind,0]) # BP
        b = np.where(mean_across_CVs_formatted_results[:,1] == excel_parameter_format[row_ind,1]) # S
        c = np.where(mean_across_CVs_formatted_results[:,2] == excel_parameter_format[row_ind,2]) # hist
        d = np.where(mean_across_CVs_formatted_results[:,3] == excel_parameter_format[row_ind,3]) # enc
        
        a1 = np.intersect1d(a,b)
        a2 = np.intersect1d(a1,c)
        a3 = np.intersect1d(a2,d)
    
        
        # Import the results into the excel spreadsheet
        if np.size(a3) != 0:
            if approx_or_no_sort == 'no-sort':
                sheet["N"+ str(row_ind+first_row)] = mean_across_CVs_formatted_results[a3,-1][0]
            elif approx_or_no_sort == 'approx-sort':
                sheet["M"+ str(row_ind+first_row)] = mean_across_CVs_formatted_results[a3,-1][0]
        else:
            if approx_or_no_sort == 'no-sort':
                sheet["N"+ str(row_ind+first_row)] = "nan"
            elif approx_or_no_sort == 'approx-sort':
                sheet["M"+ str(row_ind+first_row)] = "nan"
        
workbook.save(filename=excel_path)
        
    
 
"""
BDP S
Same thing as above, with the BDP results for each BP and S.

Basically we access the condensed BDP results from "analyse_BDP_S_pkl.py",
and format them according to the excel spreadsheet, and import the results to 
t3he excel spreadsheet.
"""

# Iterate through both BDP datasets, sum the BDP results as a function of S and BP
for dataset_count, Sabes_or_Flint in enumerate(['Flint','Sabes']):

    # Load BDP (hyper-)parameter optimised results, both validation and testing results
    with open(BDP_results_directory + '\S_vs_BDP_train_'+Sabes_or_Flint + '.pkl', 'rb') as file:
        
        
        results = pickle.load(file)
        x = results['best_test_val_params']
        x[x==0] = float('NaN')

        x2 = results['best_val_val_params']
        x2[x2==0] = float('NaN')
 
        if dataset_count == 0:
            best_test_val_res = np.nanmean(x,axis=2)
            best_val_val_res = np.nanmean(x2,axis=2)
        else:
            best_test_val_res += np.nanmean(x,axis=2)
            best_val_val_res += np.nanmean(x2,axis=2)

        
# Add zeros, so S is index instead of max val.
best_test_val_res = np.hstack((np.zeros((len(bin_vector),1)),best_test_val_res))
best_val_val_res = np.hstack((np.zeros((len(bin_vector),1)),best_val_val_res))

# Results averaged across 2 datasets
best_test_val_res = best_test_val_res / 2
best_val_val_res = best_val_val_res / 2  
all_best_res = (best_test_val_res + best_val_val_res)/2 # combined val and test results


# Rename, select which results we use (in this case: test results)
# We also have the option to analyse the validation results, but it is overfit.
res = best_test_val_res

# Format BDP data as 2D matrix with all params
formatted_results_BDP = []
for S in S_vector:
            
    # Load all possible SCLVs, just so we can see how many encoders there are per S value
    file_name = 'Stored_SCLVs_S_'+str(int(S))+'.pkl'
    with open(SCLV_directory + '\\' + file_name, 'rb') as file:      
        SCLVs = pickle.load(file)
        nb_SCLVs = len(SCLVs)
        SCLVs = np.array(SCLVs,dtype=object)
        
    # How many encoders
    nb_enc_vec = np.arange(1,len(SCLVs)+1)
    
    for BP_count, BP in enumerate(bin_vector):
        for hist_size in bits_per_channel_for_histogram_vector:
            hist = int(np.log2(hist_size))
            for nb_enc in nb_enc_vec:
                
                # Format all results together
                results = [BP,S,hist,nb_enc,res[BP_count,S]]
                formatted_results_BDP.append(results)


# Iterate through excel spreadsheet, match excel rows with "formatted_results_BDP" rows.
formatted_results_BDP  = np.array(formatted_results_BDP)
for row_ind in np.arange(len(excel_parameter_format[:,0])):

    a = np.where(formatted_results_BDP[:,0] == excel_parameter_format[row_ind,0]) # BP
    b = np.where(formatted_results_BDP[:,1] == excel_parameter_format[row_ind,1]) # S
    c = np.where(formatted_results_BDP[:,2] == excel_parameter_format[row_ind,2]) # hist
    d = np.where(formatted_results_BDP[:,3] == excel_parameter_format[row_ind,3]) # enc
    
    a1 = np.intersect1d(a,b)
    a2 = np.intersect1d(a1,c)
    a3 = np.intersect1d(a2,d)

    # Import the results into the excel spreadsheet
    if np.size(a3) != 0:
        sheet["E"+ str(row_ind+first_row)] = formatted_results_BDP[a3,-1][0]
    else:
        sheet["E"+ str(row_ind+first_row)] = "nan"
        
# Update excel spreadsheet
workbook.save(filename=excel_path)
    
